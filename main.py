from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side

# Criação de um workbook que contara todas as planilhas
arquivo_excel = Workbook()

# Pegando a planilha padrão e alterando o titulo
planilha1 = arquivo_excel.active
planilha1.title = "Gastos"

planilha2 = arquivo_excel.create_sheet("Ganhos")

print(arquivo_excel.sheetnames)
planilha1.merge_cells('B4:I4')
bd = Side(style='thick', color="FF0000")
 
top_left_cell = planilha1['B4']
top_left_cell.value = 'Fluxo de Caixa - Exercício 2018'
top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
planilha1['B9'] = ''
planilha1['B10'] = 'Vendas'
planilha1['B11'] = 'Pagamento à vista'
planilha1['B12'] = 'Pagamento depois de 30 dias'
planilha1['B13'] = 'Pagamentos depois de 60 dias'
planilha1['B14'] = 'TOTALIZADOR DE ENTRADAS'


valores = [
    ("Categoria", "Valor"),
    ("Restaurante", 45.99),
    ("Transporte", 208.45),
    ("Viagem", 558.54)
]
for linha in valores:
    planilha1.append(linha)

arquivo_excel.save("relatorio.xlsx")